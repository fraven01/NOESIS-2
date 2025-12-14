"""Excel (.xlsx and .xls) parser for document ingestion."""

from __future__ import annotations

from io import BytesIO
from typing import List, Mapping, Optional

from documents.parsers import (
    ParsedResult,
    ParsedTextBlock,
    build_parsed_result,
    build_parsed_text_block,
)
from documents.payloads import extract_payload

_MAX_SECTION_SEGMENT_LENGTH = 128


class XlsxDocumentParser:
    """Parser for Excel spreadsheet files (.xlsx and .xls)."""

    # Modern Excel format (OpenXML)
    _XLSX_MEDIA_TYPES = frozenset(
        {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
    )

    # Legacy Excel format (BIFF)
    _XLS_MEDIA_TYPES = frozenset(
        {
            "application/vnd.ms-excel",
        }
    )

    _SUPPORTED_MEDIA_TYPES = _XLSX_MEDIA_TYPES | _XLS_MEDIA_TYPES

    @staticmethod
    def _normalized_media_type(value: object) -> Optional[str]:
        if not isinstance(value, str):
            return None
        candidate = value.split(";")[0].strip().lower()
        return candidate or None

    def _infer_media_type(self, document: object) -> Optional[str]:
        blob = getattr(document, "blob", None)
        media_type = self._normalized_media_type(getattr(blob, "media_type", None))
        if media_type:
            return media_type
        meta = getattr(document, "meta", None)
        external_ref = getattr(meta, "external_ref", None)
        if isinstance(external_ref, Mapping):
            media_type = self._normalized_media_type(external_ref.get("media_type"))
            if media_type:
                return media_type
        candidate = self._normalized_media_type(getattr(document, "media_type", None))
        if candidate:
            return candidate
        return None

    def can_handle(self, document: object) -> bool:
        media_type = self._infer_media_type(document)
        return media_type in self._SUPPORTED_MEDIA_TYPES

    def _parse_xlsx(self, payload: bytes) -> tuple[List[ParsedTextBlock], dict]:
        """Parse modern .xlsx files using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("xlsx_parser_requires_openpyxl") from exc

        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)

        blocks: List[ParsedTextBlock] = []
        total_rows = 0
        total_cells = 0
        sheet_count = len(workbook.sheetnames)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            section_title = (
                sheet_name[:_MAX_SECTION_SEGMENT_LENGTH] if sheet_name else "Sheet"
            )
            section_path = (section_title,)

            sheet_content_lines = []
            for row in sheet.iter_rows(values_only=True):
                non_empty_cells = [str(cell) for cell in row if cell is not None]
                if non_empty_cells:
                    total_rows += 1
                    total_cells += len(non_empty_cells)
                    line = "\t".join(non_empty_cells)
                    sheet_content_lines.append(line)

            if sheet_content_lines:
                sheet_text = "\n".join(sheet_content_lines)
                blocks.append(
                    build_parsed_text_block(
                        text=sheet_text,
                        kind="table_summary",
                        section_path=section_path,
                    )
                )

        workbook.close()

        statistics = {
            "parser.kind": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "parser.sheets": sheet_count,
            "parser.rows": total_rows,
            "parser.cells": total_cells,
            "parser.characters": sum(len(b.text) for b in blocks),
        }

        return blocks, statistics

    def _parse_xls(self, payload: bytes) -> tuple[List[ParsedTextBlock], dict]:
        """Parse legacy .xls files using xlrd."""
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("xls_parser_requires_xlrd") from exc

        workbook = xlrd.open_workbook(file_contents=payload)

        blocks: List[ParsedTextBlock] = []
        total_rows = 0
        total_cells = 0
        sheet_count = workbook.nsheets

        for sheet_idx in range(sheet_count):
            sheet = workbook.sheet_by_index(sheet_idx)
            section_title = (
                sheet.name[:_MAX_SECTION_SEGMENT_LENGTH] if sheet.name else "Sheet"
            )
            section_path = (section_title,)

            sheet_content_lines = []
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                non_empty_cells = [str(cell) for cell in row if cell not in (None, "")]
                if non_empty_cells:
                    total_rows += 1
                    total_cells += len(non_empty_cells)
                    line = "\t".join(non_empty_cells)
                    sheet_content_lines.append(line)

            if sheet_content_lines:
                sheet_text = "\n".join(sheet_content_lines)
                blocks.append(
                    build_parsed_text_block(
                        text=sheet_text,
                        kind="table_summary",
                        section_path=section_path,
                    )
                )

        statistics = {
            "parser.kind": "application/vnd.ms-excel",
            "parser.sheets": sheet_count,
            "parser.rows": total_rows,
            "parser.cells": total_cells,
            "parser.characters": sum(len(b.text) for b in blocks),
        }

        return blocks, statistics

    def parse(self, document: object, config: object) -> ParsedResult:  # noqa: ARG002
        blob = getattr(document, "blob", None)
        payload = extract_payload(blob)
        if not payload:
            raise ValueError("xlsx_payload_missing")

        media_type = self._infer_media_type(document)

        # Route to appropriate parser based on media type
        if media_type in self._XLS_MEDIA_TYPES:
            blocks, statistics = self._parse_xls(payload)
        else:
            # Default to xlsx for modern format or unknown
            blocks, statistics = self._parse_xlsx(payload)

        return build_parsed_result(text_blocks=tuple(blocks), statistics=statistics)


__all__ = ["XlsxDocumentParser"]
